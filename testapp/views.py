from django.shortcuts import render
# Create your views here.
from testapp.models import Report,Alpha
from testapp.forms import ReportForm
import openpyxl
from pycel import ExcelCompiler
from tempfile import NamedTemporaryFile




def input(request):
    alphabet = Alpha.objects.all()
    form = ReportForm()
    if request.method == 'POST':
        form = ReportForm(request.POST,request.FILES)
        if form.is_valid():
           form.save()

    return render(request,'testapp/forms.html',{'form':form,'alphabet':alphabet})



# def index1(request):
#     file_name = request.POST.get('file_name')
#     sheet_name = request.POST.get('sheet_name')
#     min_row = request.POST.get('min_row')
#     max_row = request.POST.get('max_row')
#     column = request.POST.get('column')
#     first_cell_index = request.POST.get('first_cell_index')
#     second_cell_index = request.POST.get('second_cell_index')
#     print(type(min_row))
#     second_cell_index =  request.POST.get('second_cell_index')
#     print(type(max_row))
#     print(type(first_cell_index))
#     print(type(second_cell_index))
#     return render(request, 'testapp/index.html', {'file_name': file_name, 'sheet_name': sheet_name, 'min_row': min_row, 'max_row': max_row,  'column': column,'first_cell_index':first_cell_index,'second_cell_index':second_cell_index })
#


def index(request):
    if "GET" == request.method:
        return render(request, 'testapp/index.html', {})

    else:

        file_name = request.POST.get('file_name')
        inputsheet = request.POST.get('sheet_name')
        min_row = request.POST.get('min_row')
        max_row = request.POST.get('max_row')
        column = request.POST.get('column')
        first_cell_index = request.POST.get('first_cell_index')
        second_cell_index = request.POST.get('second_cell_index')
        inputfile  = request.FILES['file']


        con = Report(column = column, file = inputfile)
        con.save()

        sinputfile = str(inputfile)
        print(sinputfile)
        sss= sinputfile.replace(' ','_')
        print(sss)


        excel = ExcelCompiler(filename = sss)
        # print(" P16 is {}".format(excel.evaluate('Backtest!P16')))
        #
        # excel.validate_calcs(output_addrs=['Backtest!P16'])
        # print(" P16 is {}".format(excel.evaluate('Backtest!P16')))


        # excel = ExcelCompiler(filename = sinputfile)
        #
        # val = 'Backtest!'+'L16'
        # print(excel.validate_calcs(output_addrs=[val]))
        # a=excel.evaluate(val)
        # print('a',a)

        # print(" P16 is {}".format(excel.evaluate('Backtest!P16')))



        wb = openpyxl.load_workbook(filename = inputfile, data_only= True)
        # getting a particular sheet by name out of many sheets
        worksheet = wb['Backtest']



        # # # a = worksheet['E4'].value
        # # # print('a', a)
        # # # temp = worksheet['E5'].value
        # # # print('temp', temp)
        # # # result = ((temp - a) / a) * 100
        # # # print('result', result)
        # # # # l1 = []
        # # #
        # # #
        # # # for col in worksheet.iter_cols(min_col=5, max_col=5, min_row=6, max_row=203):
        # # #     for cell in col:
        # # #         new_result = ((cell.value - temp) / temp) * 100
        # # #         print('cell value', cell.value)
        # # #         print('temp', temp)
        # # #         print('new_result', new_result)
        # # #         difference = new_result - result
        # # #         print('difference', difference)
        # # #         if difference < -10:
        # # #             print('There is problem with the unique visitor data ', cell.value)
        # # #             l1.append(cell.row)
        # # #
        # # #         result = new_result
        # # #         print('result', result)
        # # #         temp = cell.value
        # # #         print('temp', temp)
        # # #     print(l1)
        # # #
        # # #  a = worksheet['N4'].value
        # # #  print('a', a)
        # # #  temp = worksheet['N5'].value
        # # #  print('temp', temp)
        # # #  result = ((temp - a) / a) * 100
        # # #  print('result', result)
        # # #  l2 =[]



        # # #  for col in worksheet.iter_cols(min_col=14, max_col=14, min_row=6, max_row=22):
        # # # #      for cell in col:
        # # # #          new_result = ((cell.value - temp) / temp) * 100
        # # # #          print('cell value', cell.value)
        # # # #          print('temp', temp)
        # # # #          print('new_result', new_result)
        # # # #          difference = new_result - result
        # # # #          print('difference', difference)
        # # # #          if difference > 10:
        # # # #              print('There is problem with the market place requests data ', cell.value)
        # # # #              l2.append(cell.row)
        # # # #
        # # # #          result = new_result
        # # # #          print('result', result)
        # # # #          temp = cell.value
        # # # #          print('temp', temp)
        # # #
        # # #  # a = worksheet['V4'].value
        # # #  # print('a', a)
        # # #  # temp = worksheet['V5'].value
        # # #  # print('temp', temp)
        # # #  # result = ((temp - a) / a) * 100
        # # #  # print('result', result)
        # # #
        # # #  # for col in worksheet.iter_cols(min_col= 22, max_col=22, min_row=6, max_row=21):
        # # #  #     for cell in col:
        # # #  #         new_result = ((cell.value - temp) / temp) * 100
        # # #  #         print('cell value', cell.value)
        # # #  #         print('temp', temp)
        # # #  #
        # # #  #         print('new_result', new_result)
        # # #  #
        # # #  #         difference = new_result - result
        # # #  #         print('difference', difference)
        # # #  #         if difference > 10:
        # # #  #             print('There is problem with the market place  revenue data ', cell.value)
        # # #  #         result = new_result
        # # #  #         print('result', result)
        # # #  #         temp = cell.value
        # # #  #         print('temp', temp)
        # # #


        icolumn = int(column)


        for col in  worksheet.iter_cols(min_col= icolumn, max_col = icolumn, min_row =4):
            for cell in col:
                if cell.value is None:
                    continue


                c= cell.coordinate
                sc = str(c)
                val = 'Backtest!' + sc
                excel.validate_calcs(output_addrs=[val])
                if excel.evaluate(val) != 0 :
                    temp =  excel.evaluate(val)
                    new_row = cell.row
                    break



        print('temp',temp)
        print('new_row',new_row)
        print(type(new_row))

        irow = new_row+1
        print(irow)
        l1 = []



        for col in  worksheet.iter_cols(min_col= icolumn, max_col = icolumn, min_row=irow ):
            for cell in col:
                c = cell.coordinate
                print(c)
                sc = str(c)
                print(sc)
                val = 'Backtest!' + sc
                print(val)
                excel.validate_calcs(output_addrs=[val])
                p = excel.evaluate(val)
                if cell.value is None:
                    continue
                difference = p - temp
                if difference > 0.1 or difference < - 0.1:
                    print('warning')
                    l1.append(cell.row)

                temp = excel.evaluate(val)



        # excel_data = list()
        # iterating over the rows and column
        # getting value from each cell in row


        # for col in worksheet.iter_cols(min_col=16,max_col=16,min_row=10,max_row=20):
        #     row_data = list()
        #
        #
        #     for cell in col:
        #         if cell.value is None:
        #             continue
        #         c = cell.coordinate
        #         sc = str(c)
        #         val = 'Backtest!' + sc
        #         excel.validate_calcs(output_addrs=[val])
        #         a = excel.evaluate(val)
        #
        #
        #
        #         row_data.append(a)
        #     excel_data.append(row_data)




        # # a = worksheet[first_cell_index].value
        # # print('a', a)
        # # temp = worksheet[second_cell_index].value
        # # print('temp', temp)
        # # result = ((temp - a) / a) * 100
        # # print('result', result)
        # # l1 = []
        # #
        # # mirow = int(min_row)
        # # marow = int(max_row)
        # # icol = int(column)
        # #
        # #
        # #
        # # for col in worksheet.iter_cols(min_col= icol, max_col= icol, min_row= mirow, max_row= marow):
        # #     for cell in col:
        # #         new_result = ((cell.value - temp) / temp) * 100
        # #         print('cell value', cell.value)
        # #         print('temp', temp)
        # #         print('new_result', new_result)
        # #         difference = new_result - result
        # #         print('difference', difference)
        # #         if difference > 10 or difference < -10:
        # #             print('There is problem with the unique visitor data ', cell.value)
        # #             l1.append(cell.row)
        # #
        # #         result = new_result
        # #         print('result', result)
        # #         temp = cell.value
        # #         print('temp', temp)
        # #     print(l1)
        #
        # #
        # # icolumn = int(column)
        # # imin_row = int(min_row)
        # # imax_row = int(max_row)
        # #
        # # temp = worksheet[first_cell_index].value * 100
        # # print('temp', temp)
        # # l1=[]
        # #
        # #
        # # for col in worksheet.iter_cols(min_col= icolumn,max_col= icolumn,min_row= imin_row,max_row= imax_row):
        # #     for cell in col:
        # #         a = cell.value *100
        # #         difference = (a - temp)
        # #         if difference >10 or difference < -10:
        # #             print(cell.row)
        # #             print('there is faulty data ',  cell.value)
        # #             l1.append(cell.row)
        # #
        # #         temp = cell.value * 100
        # #
        # #
        # # temp = worksheet['N10'].value
        # # t=temp *100
        # # print('t', t)
        #
        # print(worksheet['P10'].value)
        # wb.save()

        qs = Report.objects.all()
        qs.delete()

        return render(request, 'testapp/index.html',{'l1':l1})


